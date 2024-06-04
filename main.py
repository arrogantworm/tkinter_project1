import tkinter as tk
from tkinter import ttk
import openpyxl


# class TreeviewEdit
class TreeviewEdit(ttk.Treeview):
    def __init__(self, master, **kw):
        super().__init__(master, **kw)

        self.bind('<Double-1>', self.on_double_click)

    def on_double_click(self, event):

        region_clicked = self.identify_region(event.x, event.y)

        if region_clicked != 'cell':
            return

        column = int(self.identify_column(event.x)[1:])-1
        row = self.identify_row(event.y)
        selected_iid = self.focus()
        selected_value = self.item(selected_iid)
        item_index = self.index(self.focus()) + 2

        column_box = self.bbox(selected_iid, column)

        if column == 0:
            entry_edit = ttk.Entry(tree_frame)
            entry_edit.insert(0, string=selected_value.get('values')[0])
            entry_edit.editing_column_index = column
            entry_edit.editing_item_iid = selected_iid
            entry_edit.select_range(0, tk.END)
            entry_edit.cell_number = f'A{item_index}'
            entry_edit.focus()
            entry_edit.place(x=column_box[0],
                             y=column_box[1],
                             w=column_box[2],
                             h=column_box[3]+10)

            entry_edit.bind('<FocusOut>', self.on_focus_out)
            entry_edit.bind('<Return>', self.on_focus_out)

        if column == 1:
            age_edit = ttk.Spinbox(tree_frame, from_=18, to=100)
            age_edit.insert(0, string=selected_value.get('values')[1])
            age_edit.editing_column_index = column
            age_edit.editing_item_iid = selected_iid
            age_edit.cell_number = f'B{item_index}'
            age_edit.focus()
            age_edit.place(x=column_box[0],
                           y=column_box[1],
                           w=column_box[2] + 30,
                           h=column_box[3] + 10)

            age_edit.bind('<FocusOut>', self.on_focus_out)
            age_edit.bind('<Return>', self.on_focus_out)

        if column == 2:
            combo_list = ['Подписан', 'Не подписан', 'Другое']

            status_edit = ttk.Combobox(tree_frame, values=combo_list)
            status_edit.set(selected_value.get('values')[2])
            status_edit.editing_column_index = column
            status_edit.editing_item_iid = selected_iid
            status_edit.cell_number = f'C{item_index}'
            status_edit.focus()
            status_edit.place(x=column_box[0],
                              y=column_box[1],
                              w=column_box[2] + 20,
                              h=column_box[3] + 10)

            status_edit.bind('<FocusOut>', self.on_focus_out)
            status_edit.bind('<Return>', self.on_focus_out)

        if column == 3:
            evalue = True if selected_value.get('values')[3] == 'Трудоустроен' else False
            employed_edit = tk.BooleanVar()
            employed_edit.set(evalue)
            checkbutton_edit = ttk.Checkbutton(tree_frame, text='Трудоустроен', variable=employed_edit)
            checkbutton_edit.editing_column_index = column
            checkbutton_edit.editing_item_iid = selected_iid
            checkbutton_edit.var = employed_edit
            checkbutton_edit.cell_number = f'D{item_index}'
            checkbutton_edit.focus()
            checkbutton_edit.place(x=column_box[0],
                                   y=column_box[1],
                                   w=column_box[2] + 20,
                                   h=column_box[3] + 10)

            checkbutton_edit.bind('<FocusOut>', self.on_focus_out)
            checkbutton_edit.bind('<Return>', self.on_focus_out)

    def on_focus_out(self, event):

        selected_iid = event.widget.editing_item_iid
        column_index = event.widget.editing_column_index
        old_values = self.item(selected_iid).get('values')
        if event.widget.editing_column_index == 0:
            new_value = event.widget.get()
            new_values = old_values.copy()
            new_values[column_index] = new_value
            if old_values != new_values:

                # Insert new data in tkinter
                self.item(selected_iid, values=new_values)

                # Insert new data in excel file
                path = 'people.xlsx'
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                sheet[event.widget.cell_number] = new_value
                workbook.save(path)

        elif event.widget.editing_column_index == 1:
            new_value = int(event.widget.get())
            new_values = old_values.copy()
            new_values[column_index] = new_value

            if old_values != new_values:
                # Insert new data in tkinter
                self.item(selected_iid, values=new_values)

                # Insert new data in excel file
                path = 'people.xlsx'
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                sheet[event.widget.cell_number] = new_value
                workbook.save(path)

        elif event.widget.editing_column_index == 2:
            new_value = event.widget.get()
            new_values = old_values.copy()
            new_values[column_index] = new_value

            if old_values != new_values:
                # Insert new data in tkinter
                self.item(selected_iid, values=new_values)

                # Insert new data in excel file
                path = 'people.xlsx'
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                sheet[event.widget.cell_number] = new_value
                workbook.save(path)

        elif event.widget.editing_column_index == 3:
            new_value = event.widget.var.get()
            if new_value:
                new_value = 'Трудоустроен'
            else:
                new_value = 'Безработный'
            new_values = old_values.copy()
            new_values[column_index] = new_value

            if old_values != new_values:
                # Insert new data in tkinter
                self.item(selected_iid, values=new_values)

                # Insert new data in excel file
                path = 'people.xlsx'
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                sheet[event.widget.cell_number] = new_value
                workbook.save(path)

        event.widget.destroy()


# class AutoScrollbar
class AutoScrollbar(ttk.Scrollbar):
    def set(self, low, high):
        if float(low) <= 0.0 and float(high) >= 1.0:
            self.tk.call("grid", "remove", self)
        else:
            self.grid()
        ttk.Scrollbar.set(self, low, high)

    def pack(self, **kw):
        raise (tk.TclError, "pack cannot be used with \
               this widget")

    def place(self, **kw):
        raise (tk.TclError, "place cannot be used  with \
               this widget")


# NewHuman popup window
def new_human():

    # Добавление новой строки в excel
    def insert_row():
        name = name_entry.get()
        age = int(age_spinbox.get())
        subscription_status = status_combobox.get()
        employment_status = 'Трудоустроен' if employed.get() else 'Не трудоустроен'

        # Insert row into excel sheet
        path = 'people.xlsx'
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [name, age, subscription_status, employment_status]
        sheet.append(row_values)
        workbook.save(path)

        # Insert row into treeview
        treeview.insert('', tk.END, values=row_values)

        # Clear form
        name_entry.delete(0, 'end')
        name_entry.insert(0, 'Имя')
        age_spinbox.delete(0, 'end')
        age_spinbox.insert(0, 'Возраст')
        status_combobox.set(combo_list[0])
        checkbutton.state(['!selected'])

        # Close form
        neworderwin.destroy()
        neworderwin.update()

        # Scroll to the bottom
        treeview.yview_moveto(1)

    # NewHuman popup window
    neworderwin = tk.Toplevel(root)
    neworderwin.title("Добавить")
    neworderwin.resizable(False, False)
    w = neworderwin.winfo_screenwidth()
    h = neworderwin.winfo_screenheight()
    w = w // 2 - 100
    h = h // 2 - 200
    neworderwin.geometry(f'+{w}+{h}')

    no_frame = ttk.Frame(neworderwin)
    no_frame.pack()

    form_frame = ttk.LabelFrame(no_frame, text='Добавить')
    form_frame.grid(row=0, column=0, padx=20, pady=10)

    name_entry = ttk.Entry(form_frame)
    name_entry.insert(0, 'Имя')
    name_entry.bind("<FocusIn>", lambda args: name_entry.delete('0', 'end'))
    name_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky='ew')

    age_spinbox = ttk.Spinbox(form_frame, from_=18, to=100)
    age_spinbox.insert(0, 'Возраст')
    age_spinbox.bind("<FocusIn>", lambda args: age_spinbox.delete('0', 'end'))
    age_spinbox.grid(row=1, column=0, padx=5, pady=5, sticky='ew')

    combo_list = ['Подписан', 'Не подписан', 'Другое']

    status_combobox = ttk.Combobox(form_frame, values=combo_list)
    status_combobox.current(0)
    status_combobox.grid(row=2, column=0, padx=5, pady=5, sticky='ew')

    employed = tk.BooleanVar()
    checkbutton = ttk.Checkbutton(form_frame, text='Трудоустроен', variable=employed)
    checkbutton.grid(row=3, column=0, padx=5, pady=5, sticky='nsew')

    button = ttk.Button(form_frame, text='Добавить', command=insert_row)
    button.grid(row=4, column=0, padx=5, pady=5, sticky='nsew')


# Theme switch
def toggle_mode():
    if mode_switch.get():
        style.theme_use('forest-light')
    else:
        style.theme_use('forest-dark')


# Load Data
def load_data():
    path = 'people.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)

    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)


# Initialize
root = tk.Tk()
root.resizable(True, True)
root.title('tkinter обучение')
w = root.winfo_screenwidth()
h = root.winfo_screenheight()
w = w // 2 - 200
h = h // 2 - 200
root.geometry(f'+{w}+{h}')


# Style
style = ttk.Style(root)
root.tk.call('source', 'forest-light.tcl')
root.tk.call('source', 'forest-dark.tcl')
style.theme_use('forest-dark')

# Top Menu
menu = tk.Menu(root, tearoff=False)
root.config(menu=menu)
menu.add_command(label='Добавить', command=new_human)
mode_switch = tk.BooleanVar()
menu.add_checkbutton(label='Тема', variable=mode_switch, onvalue=True, offvalue=False, command=toggle_mode)


# TreeView
tree_frame = ttk.Frame(root)
tree_frame.pack(fill='both', expand=True)
tree_frame.grid_rowconfigure(0, weight=1)
tree_frame.grid_columnconfigure(0, weight=1)
tree_scroll_v = AutoScrollbar(tree_frame, orient='vertical')
tree_scroll_v.grid(row=0, column=1, sticky='nse')
tree_scroll_h = AutoScrollbar(tree_frame, orient='horizontal')
tree_scroll_h.grid(row=1, column=0, sticky='sew')

cols = ('Имя', 'Возраст', 'Подписка', 'Трудоустройство')
treeview = TreeviewEdit(tree_frame,
                        show='headings',
                        yscrollcommand=tree_scroll_v.set,
                        xscrollcommand=tree_scroll_h.set,
                        columns=cols)
treeview.column('Имя', width=150, minwidth=150)
treeview.column('Возраст', width=50, minwidth=50)
treeview.column('Подписка', width=100, minwidth=100)
treeview.column('Трудоустройство', width=100, minwidth=100)
treeview.grid(row=0, column=0, sticky='nsew')
tree_scroll_v.config(command=treeview.yview)
tree_scroll_h.config(command=treeview.xview)
load_data()


root.mainloop()
